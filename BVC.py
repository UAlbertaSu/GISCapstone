import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font
import requests
from datetime import datetime, timedelta, timezone
import time
import json
import os
import urllib.parse
from dotenv import load_dotenv

class BVCApp:
    def __init__(self, root):
        self.root = root
        self.root.title("BVC Student-Host Matching System")
        self.root.geometry("600x400")
        
        # Variables
        self.student_file = tk.StringVar()
        self.host_file = tk.StringVar()
        self.output_folder = tk.StringVar()
        
        # GUI Elements
        self.create_widgets()
    
    def create_widgets(self):
        # Frame for inputs
        input_frame = ttk.LabelFrame(self.root, text="Input Files", padding=10)
        input_frame.pack(fill="both", expand=True, padx=10, pady=10)
        
        # Student File Selection
        ttk.Label(input_frame, text="Student Classlist:").grid(row=0, column=0, sticky="w")
        ttk.Entry(input_frame, textvariable=self.student_file, width=50).grid(row=0, column=1, padx=5)
        ttk.Button(input_frame, text="Browse", command=lambda: self.browse_file(self.student_file)).grid(row=0, column=2)
        
        # Host File Selection
        ttk.Label(input_frame, text="Host Locations:").grid(row=1, column=0, sticky="w")
        ttk.Entry(input_frame, textvariable=self.host_file, width=50).grid(row=1, column=1, padx=5)
        ttk.Button(input_frame, text="Browse", command=lambda: self.browse_file(self.host_file)).grid(row=1, column=2)
        
        # Output Folder Selection
        ttk.Label(input_frame, text="Output Folder:").grid(row=2, column=0, sticky="w")
        ttk.Entry(input_frame, textvariable=self.output_folder, width=50).grid(row=2, column=1, padx=5)
        ttk.Button(input_frame, text="Browse", command=lambda: self.browse_folder(self.output_folder)).grid(row=2, column=2)
        
        # Run Button
        run_button = ttk.Button(self.root, text="Run Matching", command=self.run_matching)
        run_button.pack(pady=10)
        
        # Progress Bar
        self.progress = ttk.Progressbar(self.root, orient="horizontal", length=400, mode="determinate")
        self.progress.pack(pady=10)
        
        # Status Label
        self.status = ttk.Label(self.root, text="Ready")
        self.status.pack()
    
    def browse_file(self, file_var):
        filename = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
        if filename:
            file_var.set(filename)
    
    def browse_folder(self, folder_var):
        foldername = filedialog.askdirectory()
        if foldername:
            folder_var.set(foldername)
    
    def update_status(self, message):
        self.status.config(text=message)
        self.root.update_idletasks()
    
    def run_matching(self):
        # Validate inputs
        if not all([self.student_file.get(), self.host_file.get(), self.output_folder.get()]):
            messagebox.showerror("Error", "Please select all required files and output folder.")
            return
        
        try:
            self.update_status("Processing...")
            self.progress["value"] = 0
            
            # Run the matching process
            self.main(
                self.student_file.get(),
                self.host_file.get(),
                self.output_folder.get()
            )
            
            self.progress["value"] = 100
            self.update_status("Completed successfully!")
            messagebox.showinfo("Success", "Matching completed. Output file generated.")
        
        except Exception as e:
            self.update_status(f"Error: {str(e)}")
            messagebox.showerror("Error", f"An error occurred: {str(e)}")
    
    # ===== Original Functions (Modified for GUI) ===== #
    def get_next_monday_8am(self):
        now = datetime.now(timezone.utc)
        days_until_monday = (7 - now.weekday()) % 7
        next_monday = now + timedelta(days=days_until_monday)
        nextMonday_MT = next_monday.astimezone()
        return nextMonday_MT.replace(hour=8, minute=0, second=0, microsecond=0).isoformat()

    def get_distance(self, origin, destination):
        load_dotenv()
        api_key = os.getenv("API_KEY")
        url = "https://routes.googleapis.com/distanceMatrix/v2:computeRouteMatrix"
        headers = {
            "Content-Type": "application/json",
            "X-Goog-Api-Key": api_key,
            "X-Goog-FieldMask": "originIndex,destinationIndex,duration,distanceMeters,status"
        }
        payload = {
            "origins": [{"waypoint": {"address": origin}}],
            "destinations": [{"waypoint": {"address": address}} for address in destination],
            "travelMode": "TRANSIT",
            "transitPreferences": {"allowedTravelModes": ["BUS", "RAIL"]},
            "departureTime": self.get_next_monday_8am()
        }
        response = requests.post(url, headers=headers, data=json.dumps(payload))
        return response.json()

    def processSheet(self, wb, dict):
        sheet = wb.active
        for row_num, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            if row_num == 1:
                field = row
            else:
                if field[0] == "Student ID#":
                    studentID = int(round(row[0]))
                    dict[studentID] = row[1:3]
                else:
                    hostList = list(row[:2])
                    hostList.append(row[5])
                    hostList.append(row[8])
                    dict[row[3]] = hostList
        return dict

    def main(self, studentFilePath, practicumLocationPath, outputPath):
        timeStamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        os.makedirs(outputPath, exist_ok=True)

        studentWb = load_workbook(studentFilePath)
        hostListWb = load_workbook(practicumLocationPath)
        
        studentDict = {}
        hostDict = {}
        travelTimeMatrix = {}

        self.update_status("Processing student data...")
        studentDict = self.processSheet(studentWb, studentDict)
        
        self.update_status("Processing host data...")
        hostDict = self.processSheet(hostListWb, hostDict)
        
        self.update_status("Calculating travel times...")
        total_students = len(studentDict)
        for i, (k, v) in enumerate(studentDict.items(), 1):
            origin = (v[0] + " " + v[1])
            travelTimeMatrix[k] = {}
            studentCity = origin.strip().split()[-1]
            hostDestination = []
            hostKeys = []
            
            for key, value in hostDict.items():
                destination = (key + " " + value[2])
                hostCity = destination.strip().split()[-1]
                
                if studentCity.lower() != hostCity.lower():
                    continue
                
                hostDestination.append(destination)
                hostKeys.append(key)
            
            if hostDestination:
                distance = self.get_distance(origin, hostDestination)
                timeToTravel = []
                
                for res in distance:
                    destIndex = res['destinationIndex']
                    duration = int(res['duration'].rstrip('s'))
                    
                    if duration < 3600:
                        destinationKey = hostKeys[destIndex]
                        timeToTravel.append({destinationKey: duration})
                
                if not timeToTravel:
                    shortest = min(distance, key=lambda x: int(x['duration'].rstrip('s')))
                    shortest_duration = int(shortest['duration'].rstrip('s'))
                    shortest_dest_index = shortest['destinationIndex']
                    destinationKey = hostKeys[shortest_dest_index]
                    timeToTravel.append({destinationKey: shortest_duration})
                
                travelTimeMatrix[k] = timeToTravel
            
            self.progress["value"] = (i / total_students) * 100
            self.update_status(f"Processing student {i} of {total_students}...")
        
        self.update_status("Generating output...")
        r = Workbook()
        rs = r.active
        rs.title = "BVCOutput"
        
        headers = ["Student ID", "Origin", "Destination", "Duration in minutes", "Google Maps"]
        for col, header in enumerate(headers, 1):
            cell = rs.cell(row=1, column=col, value=header)
            cell.font = Font(b=True)
        
        row_num = 2
        for student, locations in travelTimeMatrix.items():
            rs.cell(row=row_num, column=1, value=student)
            rs.cell(row=row_num, column=2, value=f"{studentDict[student][0]}, {studentDict[student][1]}")
            studentAddress = urllib.parse.urlencode({"origin": f"{studentDict[student][0]}, {studentDict[student][1]}"})
            
            for location in locations:
                for address, duration in location.items():
                    destinationAddress = urllib.parse.urlencode({"destination": address.strip()})
                    url = f"https://www.google.com/maps/dir/?api=1&{studentAddress}&{destinationAddress}&travelmode=transit"
                    
                    rs.cell(row=row_num, column=3, value=address.strip())
                    rs.cell(row=row_num, column=4, value=duration/60)
                    rs.cell(row=row_num, column=5, value=url)
                    row_num += 1
        
        for column in rs.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            rs.column_dimensions[column_letter].width = adjusted_width
        
        filename = f"BVCoutput_{timeStamp}.xlsx"
        r.save(f"{outputPath}/{filename}")

if __name__ == "__main__":
    root = tk.Tk()
    app = BVCApp(root)
    root.mainloop()