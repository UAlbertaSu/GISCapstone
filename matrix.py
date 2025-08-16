from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Font
import requests
from datetime import datetime, timedelta, timezone
import time
import json
import os
import urllib.parse 

from dotenv import load_dotenv



def get_next_monday_8am():
    now = datetime.now(timezone.utc)
    days_until_monday = (7 - now.weekday()) % 7
    next_monday = now + timedelta(days=days_until_monday)
    nextMonday_MT = next_monday.astimezone()
    return nextMonday_MT.replace(hour=8, minute=0, second=0, microsecond=0).isoformat()

def get_distance(origin, destination):
    
    load_dotenv()
    api_key = os.getenv("API_KEY")

    # API Config
    
    url = "https://routes.googleapis.com/distanceMatrix/v2:computeRouteMatrix"

    # Request Headers
    headers = {
        "Content-Type": "application/json",
        "X-Goog-Api-Key": api_key,
        "X-Goog-FieldMask": "originIndex,destinationIndex,duration,distanceMeters,status"
    }

    # Request Body with Variables
    payload = {
        "origins": [{
            "waypoint": {
                "address": origin  # Using origin variable
            }
        }],
        "destinations": [{
            "waypoint": {
                "address": address  # Using destination variable
            }
        } for address in destination],
        "travelMode": "TRANSIT",
        "transitPreferences": {
            "allowedTravelModes": ["BUS", "RAIL"]
        },
        "departureTime": get_next_monday_8am()
    }

    # Send Request
    response = requests.post(url, headers=headers, data=json.dumps(payload))
    # print(response.json())
    return response.json()

def processSheet(wb, dict):
    # parse the student ID, city and their address in dictionary
    
    sheet = wb.active
    
    
    for row_num, row in enumerate(sheet.iter_rows(values_only = True), start = 1):
        if row_num == 1:
            field = row
        else:
            if field[0] == "Student ID#":
                studentID = int(round(row[0]))
                dict[studentID] = row[1:3]
                
            else:
                hostList = list(row[:2])
                hostList.append(row[5])
                hostList.append(row[8])
                dict[row[3]] = hostList
    
    return dict
    
def main():
    # Delete this check in MVP
     
    timeStamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    while True:
        try:
            studentFilePath = input("Please enter the absolute path of the Classlist excel sheet: ")
            practicumLocationPath = input("Please enter the absolute folder path of the host location excel sheet: ")
            outputPath = input("Please give the folder location of where you want to save the output excel spreadsheet: ")
            os.makedirs(outputPath, exist_ok=True)

            studentWb = load_workbook(studentFilePath)
            hostListWb = load_workbook(practicumLocationPath)
            break
        except:
            print("Please make sure the path is correct and includes the file name and extension. (.xlsx) where applicable")
            
    studentDict = {}
    hostDict = {}
    
    # parse the origins (student address) 
    print("Student Location \n")
    studentDict = processSheet(studentWb, studentDict)
  
    # parse the destination and attributes
    hostDict = processSheet(hostListWb, hostDict)
    travelTimeMatrix = {}

    # call API
    for k,v in studentDict.items():
        #combine the City name with the address to specify search parameter for google maps
        origin = (v[0]+" "+ v[1])
        travelTimeMatrix[k] = {}
        studentCity = origin.strip().split()[-1]
        hostDestination = []
        hostKeys = []
        for key, value in hostDict.items():
            
            destination = (key+" "+value[2])
            
            hostCity = destination.strip().split()[-1]
            
            if studentCity.lower() != hostCity.lower():
                continue
            # add additional checks here to trim calls
            # for example student - host mentorship criteria
            
            hostDestination.append(destination)
            hostKeys.append(key)
        
        if hostDestination:
        
            distance = get_distance(origin, hostDestination)
            # print(distance)
            timeToTravel =[]
            for res in distance:
                
                destIndex = res['destinationIndex']
                duration = int(res['duration'].rstrip('s'))
                
                if duration < 3600:
                    destinationKey = hostKeys[destIndex]
                    timeToTravel.append({destinationKey: duration})
                    
            if not timeToTravel:
                shortest = min(distance, key=lambda x: int(x['duration'].rstrip('s')))
                shortest_duration = int(shortest['duration'].rstrip('s'))
                shortest_dest_index = shortest['destinationIndex']
                destinationKey = hostKeys[shortest_dest_index]
                timeToTravel.append({destinationKey: shortest_duration})
            
            travelTimeMatrix[k] = timeToTravel
    
    print(travelTimeMatrix)
    # print it out nicely and create a maps url 
    # Put result in Excel. 
    # r = results
    # rs = resultSheet
    r = Workbook()
    rs = r.active
    rs.title = "BVCOutput"
    
    headers = ["Student ID", "Origin", "Destination", "Duration in minutes", "Google Maps"]
    for col, header in enumerate(headers, 1):
        cell = rs.cell(row=1, column = col, value = header)
        cell.font = Font(b = True)

    row_num = 2
       
    for student, locations in travelTimeMatrix.items():
        print(f"student ID : {student}, Origin: {studentDict[student][0]}, {studentDict[student][1]}")
        rs.cell(row = row_num, column = 1, value = student)
        rs.cell(row = row_num, column= 2, value = studentDict[student][0]+","+ studentDict[student][1])
        studentAddress =  urllib.parse.urlencode({"origin" : studentDict[student][0]+","+studentDict[student][1]})
        for location in locations:
            for address, duration in location.items():
                destinationAddress =urllib.parse.urlencode({"destination" : {address.strip()} })
                url = "https://www.google.com/maps/dir/?api=1&"+studentAddress+ "&"+ destinationAddress+"&travelmode=transit"
                print(f" Address: {address.strip()} | {duration} seconds | {url}")
                rs.cell(row= row_num, column=3, value = address.strip())
                rs.cell(row= row_num, column=4, value = duration/60)
                rs.cell(row= row_num, column=5, value = url)
                row_num += 1
    
    # format the cells width
    for column in rs.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        rs.column_dimensions[column_letter].width = adjusted_width
    #Generate output excel 
    filename = f"BVCoutput_{timeStamp}.xlsx"
    r.save(f"{outputPath}\{filename}")

    print(f"Output created in {outputPath}\{filename}")
main()

